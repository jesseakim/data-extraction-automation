"""
Portfolio-safe demo version.

This file is a sanitized recreation of an internal automation pattern. It uses
generic company names, placeholder URLs, demo location codes, and sample feature
names. It is intended to demonstrate Python automation, Selenium scraping,
Excel report generation, dependency checks, authentication/session handling,
and file delivery workflows without exposing proprietary systems or data.
"""

import time
import os
import json
import pickle
import subprocess
import getpass
import shutil
import socket
from datetime import datetime, timezone


# ====================================================================
# DEPENDENCY CHECKS WITH FRIENDLY DOWNLOAD INSTRUCTIONS
# ====================================================================
def check_dependency(package_name, pip_name=None):
    if pip_name is None:
        pip_name = package_name
    try:
        __import__(package_name)
        return True
    except ImportError:
        print(f"\n{'='*60}")
        print(f"  ⚠️  Missing required package: {package_name}")
        print(f"{'='*60}")
        print(f"\n  Download: Install it using Command Prompt (CMD).\n")
        print(f"  How to open CMD:")
        print(f"    1. Press  Win + R  on your keyboard")
        print(f"    2. Type   cmd   and press Enter")
        print(f"    3. In the black window, type the command below and press Enter:\n")
        print(f"       pip install {pip_name}\n")
        print(f"    4. Wait for it to finish, then re-run this script.\n")
        print(f"{'='*60}")
        choice = input("\n  Press Enter to exit so you can install it (or type 'skip' to continue): ").strip().lower()
        if choice != "skip":
            exit()
        return False


sel_ok = check_dependency("selenium")
xl_ok = check_dependency("openpyxl")

if not sel_ok or not xl_ok:
    print("\n  ❌ Cannot continue without required packages.")
    input("\n  Press Enter to exit...")
    exit()

requests_ok = False
try:
    import requests as req_lib
    requests_ok = True
except ImportError:
    print(f"\n  ℹ️  Optional package 'requests' not installed.")
    print(f"     Cloud Document Portal API upload will be unavailable.")
    print(f"     To enable it, run:  pip install requests\n")

win32_ok = False
try:
    import win32com.client
    win32_ok = True
except ImportError:
    print(f"\n  ℹ️  Optional package 'pywin32' not installed.")
    print(f"     Email drafting will be skipped.")
    print(f"     To enable it, run:  pip install pywin32\n")

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ====================================================================
# HELPER: TIMEZONE-AWARE TIMESTAMPS
# ====================================================================
def local_now():
    return datetime.now().astimezone()


def utc_now():
    return datetime.now(timezone.utc)


def fmt_local(dt=None, fmt="%Y-%m-%d %H:%M:%S %Z"):
    if dt is None:
        dt = local_now()
    return dt.strftime(fmt)


def fmt_utc(dt=None, fmt="%Y-%m-%d %H:%M:%S UTC"):
    if dt is None:
        dt = utc_now()
    return dt.strftime(fmt)


# ====================================================================
# CUSTOM EXCEPTIONS
# ====================================================================
class NoEditPermissionError(Exception):
    pass


class VPNRequiredError(Exception):
    pass


# ====================================================================
# VPN / CSA DETECTION
# ====================================================================
CSA_SIGNATURES = [
    "corporate secure access",
    "posture cookie",
    "install or repair your csa plugin",
    "you must use corporate secure access",
    "csa plugin",
]

VPN_TEST_HOSTS = [
    ("na.delivery-config.example.internal", 443),
    ("eu.delivery-config.example.internal", 443),
]


def is_csa_blocked(driver):
    try:
        page_text = driver.find_element(By.TAG_NAME, "body").text.lower()
        matches = sum(1 for sig in CSA_SIGNATURES if sig in page_text)
        return matches >= 2
    except Exception:
        return False


def check_network_reachability():
    for host, port in VPN_TEST_HOSTS:
        try:
            addr = socket.getaddrinfo(host, port, socket.AF_INET, socket.SOCK_STREAM)
            if not addr:
                continue
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(10)
            result = sock.connect_ex((host, port))
            sock.close()
            if result == 0:
                return "reachable"
        except socket.gaierror:
            continue
        except socket.timeout:
            continue
        except Exception:
            continue
    return "unreachable"


def is_cisco_vpn_running():
    vpn_processes = ["vpnui.exe", "vpnagent.exe", "acwebhelper.exe", "vpndownloader.exe"]
    for proc in vpn_processes:
        try:
            result = subprocess.run(
                ["tasklist", "/FI", f"IMAGENAME eq {proc}"],
                capture_output=True, text=True, timeout=10,
            )
            if proc.lower() in result.stdout.lower():
                return True
        except Exception:
            pass
    return False


def find_cisco_vpn_path():
    possible_paths = [
        os.path.join(os.environ.get("ProgramFiles(x86)", ""), "Corporate VPN", "Corporate VPN Client", "vpnui.exe"),
        os.path.join(os.environ.get("ProgramFiles", ""), "Corporate VPN", "Corporate VPN Client", "vpnui.exe"),
        os.path.join(os.environ.get("ProgramFiles(x86)", ""), "Corporate VPN", "Corporate VPN Client", "vpnui.exe"),
        os.path.join(os.environ.get("ProgramFiles", ""), "Corporate VPN", "Corporate VPN Client", "vpnui.exe"),
    ]
    for path in possible_paths:
        if path and os.path.exists(path):
            return path
    if shutil.which("vpnui.exe"):
        return shutil.which("vpnui.exe")
    return None


def prompt_vpn_before_start():
    print(f"\n[SETUP] Checking network connectivity...\n")

    vpn_running = is_cisco_vpn_running()
    vpn_path = find_cisco_vpn_path()

    if vpn_running:
        print(f"  ✅ Corporate VPN Client is running")
    else:
        print(f"  ℹ️  Corporate VPN Client not detected as running")

    print(f"  🔍 Testing connection to LargeItem planning site...")
    status = check_network_reachability()

    if status == "reachable":
        print(f"  ✅ Network connection verified! Site is reachable.")
        print(f"  ℹ️  Note: The browser may still need the CSA plugin to")
        print(f"     fully access the site. If you see an CSA block page")
        print(f"     after the browser opens, the tool will guide you.\n")
        return True

    print(f"  🔒 Cannot reach the LargeItem planning site!")
    print(f"     This usually means VPN is not connected.\n")
    print(f"  {'='*56}")
    print(f"  🔒  VPN CONNECTION REQUIRED")
    print(f"  {'='*56}")
    print()
    print(f"  The LargeItem planning site is not reachable from your network.")
    print(f"  If you are working remotely, you need to connect to")
    print(f"  Corporate VPN Client (VPN) before the scraper can run.\n")
    print(f"  If you are in an Acme Retail office on the corporate network,")
    print(f"  this may indicate a network issue — try again in a moment.\n")

    if vpn_running:
        print(f"  🔍 Corporate VPN Client is running but you may not")
        print(f"     be connected. Check its status:\n")
        print(f"  Steps:")
        print(f"    1. Look for the Corporate VPN icon in your system tray")
        print(f"       (bottom-right corner of your screen, near the clock)")
        print(f"    2. Click the icon and make sure it says 'Connected'")
        print(f"    3. If disconnected, click 'Connect' and log in")
        print(f"    4. Once connected, come back here and press Enter\n")
    elif vpn_path:
        print(f"  ✅ Corporate VPN Client found on your computer.\n")
        print(f"  Steps:")
        print(f"    1. Open Corporate VPN Client")
        print(f"       (Search 'Corporate VPN' in your Start menu, or check system tray)")
        print(f"    2. Click 'Connect'")
        print(f"    3. Enter your credentials and complete MFA")
        print(f"    4. Wait until it says 'Connected'")
        print(f"    5. Come back here and press Enter\n")

        launch = input(f"  Would you like me to open Corporate VPN Client for you? (y/n): ").strip().lower()
        if launch in ("y", "yes"):
            try:
                subprocess.Popen([vpn_path])
                print(f"  ✅ Launching Corporate VPN Client...")
                print(f"     Connect to VPN, then come back here.\n")
            except Exception as e:
                print(f"  ⚠️  Could not launch automatically: {e}")
                print(f"     Please open it manually from your Start menu.\n")
    else:
        print(f"  ⚠️  Corporate VPN Client was not found on your computer.\n")
        print(f"  Steps:")
        print(f"    1. Search for 'Corporate VPN Client' or 'Corporate VPN'")
        print(f"       in your Start menu")
        print(f"    2. If not installed, contact IT Support to install it")
        print(f"    3. Once installed, open it and click 'Connect'")
        print(f"    4. Enter your credentials and complete MFA")
        print(f"    5. Wait until it says 'Connected'")
        print(f"    6. Come back here and press Enter\n")

    max_attempts = 5
    for attempt in range(1, max_attempts + 1):
        print(f"  ─────────────────────────────────────────────────")
        choice = input(f"  Press Enter once VPN is connected (or type 'quit' to stop): ").strip().lower()

        if choice in ("quit", "q", "exit", "stop"):
            print(f"\n  → Stopping. Connect to VPN and re-run the script.")
            return False

        print(f"\n  🔄 Checking connection... (attempt {attempt}/{max_attempts})")

        if is_cisco_vpn_running():
            print(f"  ✅ Corporate VPN process detected")
        else:
            print(f"  ⚠️  Corporate VPN process not detected — make sure it's running")

        status = check_network_reachability()

        if status == "reachable":
            print(f"  ✅ Network connection verified! Site is reachable.")
            print(f"  💡 The scraper will now proceed. After any required login,")
            print(f"     you can minimize and keep working on other things.\n")
            return True
        else:
            if attempt < max_attempts:
                print(f"  ❌ Site still unreachable.")
                if is_cisco_vpn_running():
                    print(f"     Corporate VPN is running — it may still be connecting.")
                    print(f"     Wait a few seconds for it to fully connect,")
                    print(f"     then press Enter to try again.\n")
                else:
                    print(f"     Corporate VPN doesn't appear to be running.")
                    print(f"     Please open it and connect before trying again.\n")
            else:
                print(f"  ❌ Cannot reach site after {max_attempts} attempts.")
                print(f"     Please troubleshoot your VPN and re-run the script.\n")
                return False

    return False


def prompt_vpn_connection_browser(driver, region_url):
    print()
    print(f"  {'='*56}")
    print(f"  🔒  VPN / CSA ISSUE DETECTED")
    print(f"  {'='*56}")
    print()
    print(f"  The browser hit the Corporate Secure Access (CSA) page.")
    print(f"  This can mean:")
    print(f"    • Your VPN connection dropped")
    print(f"    • The CSA browser plugin needs to be installed or repaired")
    print(f"    • Your CSA session expired\n")

    vpn_running = is_cisco_vpn_running()
    vpn_path = find_cisco_vpn_path()
    net_status = check_network_reachability()

    if net_status == "reachable":
        print(f"  🔍 Network is reachable — this is likely an CSA plugin issue,")
        print(f"     not a VPN disconnect.\n")
        print(f"  Steps to try:")
        print(f"    1. Close all browser windows")
        print(f"    2. Reopen your browser and go to any company internal site")
        print(f"    3. If CSA prompts you, follow its instructions")
        print(f"    4. Once you can access internal sites, come back here")
        print(f"    5. Press Enter and the scraper will retry\n")
        print(f"  If that doesn't work:")
        print(f"    • Try restarting your computer")
        print(f"    • Contact IT Support to repair the CSA plugin\n")
    elif vpn_running:
        print(f"  🔍 Corporate VPN Client is running but the site is")
        print(f"     unreachable. You may have been disconnected.\n")
        print(f"  Steps:")
        print(f"    1. Check the Corporate VPN icon in your system tray")
        print(f"    2. Make sure it says 'Connected'")
        print(f"    3. If disconnected, click 'Connect' and log in")
        print(f"    4. Once connected, come back here and press Enter\n")
    elif vpn_path:
        print(f"  Steps:")
        print(f"    1. Open Corporate VPN Client")
        print(f"    2. Click 'Connect' and enter credentials")
        print(f"    3. Wait until it says 'Connected'")
        print(f"    4. Come back here and press Enter\n")

        launch = input(f"  Open Corporate VPN Client for you? (y/n): ").strip().lower()
        if launch in ("y", "yes"):
            try:
                subprocess.Popen([vpn_path])
                print(f"  ✅ Launching... connect then come back here.\n")
            except Exception:
                print(f"  ⚠️  Could not launch. Open it manually.\n")
    else:
        print(f"  Steps:")
        print(f"    1. Open Corporate VPN Client from your Start menu")
        print(f"    2. Connect to VPN")
        print(f"    3. Come back here and press Enter\n")

    max_attempts = 5
    for attempt in range(1, max_attempts + 1):
        print(f"  ─────────────────────────────────────────────────")
        choice = input(f"  Press Enter once ready (or type 'quit' to stop): ").strip().lower()

        if choice in ("quit", "q", "exit", "stop"):
            print(f"\n  → Stopping. Fix VPN/CSA and re-run the script.")
            return False

        print(f"\n  🔄 Checking... (attempt {attempt}/{max_attempts})")

        net = check_network_reachability()
        if net == "reachable":
            print(f"  ✅ Network is reachable")
        else:
            if attempt < max_attempts:
                print(f"  ❌ Network still unreachable — VPN may not be connected")
                if is_cisco_vpn_running():
                    print(f"     Corporate VPN is running but may still be connecting.\n")
                else:
                    print(f"     Corporate VPN doesn't appear to be running.\n")
                continue
            else:
                print(f"  ❌ Network unreachable after {max_attempts} attempts.\n")
                return False

        try:
            driver.get(region_url)
            time.sleep(8)

            if is_csa_blocked(driver):
                if attempt < max_attempts:
                    print(f"  ❌ Still seeing CSA block page in the browser.")
                    print(f"     Network is reachable, so this is likely an CSA")
                    print(f"     plugin issue rather than a VPN problem.")
                    print(f"     Try closing all browser windows, reopening, and")
                    print(f"     visiting an company internal site first.\n")
                else:
                    print(f"  ❌ Still blocked after {max_attempts} attempts.")
                    print(f"     You may need to restart or contact IT Support.\n")
                    return False
            else:
                print(f"  ✅ Page loaded successfully!")
                print(f"  💡 The browser will continue on its own. Feel free to")
                print(f"     minimize and keep working on other things.\n")
                return True

        except Exception as e:
            if attempt < max_attempts:
                print(f"  ⚠️  Error: {e}. Trying again.\n")
            else:
                print(f"  ❌ Could not verify after {max_attempts} attempts.\n")
                return False

    return False


# ====================================================================
# BROWSER DETECTION & SELECTION
# ====================================================================
BROWSER_INFO = {
    "edge": {
        "name": "Microsoft Edge",
        "process": "msedge.exe",
        "driver_process": "msedgedriver.exe",
        "paths": [
            os.path.join(os.environ.get("ProgramFiles(x86)", ""), "Microsoft", "Edge", "Application", "msedge.exe"),
            os.path.join(os.environ.get("ProgramFiles", ""), "Microsoft", "Edge", "Application", "msedge.exe"),
            os.path.join(os.environ.get("LOCALAPPDATA", ""), "Microsoft", "Edge", "Application", "msedge.exe"),
        ],
    },
    "chrome": {
        "name": "Google Chrome",
        "process": "chrome.exe",
        "driver_process": "chromedriver.exe",
        "paths": [
            os.path.join(os.environ.get("ProgramFiles", ""), "Google", "Chrome", "Application", "chrome.exe"),
            os.path.join(os.environ.get("ProgramFiles(x86)", ""), "Google", "Chrome", "Application", "chrome.exe"),
            os.path.join(os.environ.get("LOCALAPPDATA", ""), "Google", "Chrome", "Application", "chrome.exe"),
        ],
    },
    "firefox": {
        "name": "Mozilla Firefox",
        "process": "firefox.exe",
        "driver_process": "geckodriver.exe",
        "paths": [
            os.path.join(os.environ.get("ProgramFiles", ""), "Mozilla Firefox", "firefox.exe"),
            os.path.join(os.environ.get("ProgramFiles(x86)", ""), "Mozilla Firefox", "firefox.exe"),
        ],
    },
}

BROWSER_PRIORITY = ["edge", "chrome", "firefox"]


def is_browser_installed(browser_key):
    info = BROWSER_INFO[browser_key]
    for path in info["paths"]:
        if path and os.path.exists(path):
            return True
    if shutil.which(info["process"]):
        return True
    return False


def is_browser_running(browser_key):
    process_name = BROWSER_INFO[browser_key]["process"]
    try:
        result = subprocess.run(
            ["tasklist", "/FI", f"IMAGENAME eq {process_name}"],
            capture_output=True, text=True, timeout=10,
        )
        return process_name.lower() in result.stdout.lower()
    except Exception:
        return False


def get_installed_browsers():
    return [b for b in BROWSER_PRIORITY if is_browser_installed(b)]


def get_running_browsers():
    return [b for b in BROWSER_PRIORITY if is_browser_running(b)]


def pick_best_browser():
    installed = get_installed_browsers()
    running = get_running_browsers()

    if not installed:
        return None, False

    available = [b for b in installed if b not in running]

    if available:
        chosen = available[0]
        print(f"  ✅ Using {BROWSER_INFO[chosen]['name']} (installed and not in use)")
        return chosen, False

    print(f"\n  ⚠️  All installed browsers are currently open:")
    for b in installed:
        print(f"      • {BROWSER_INFO[b]['name']}")

    print(f"\n  The scraper needs to briefly close one browser to run.")
    print(f"  Options:")
    for i, b in enumerate(installed, 1):
        print(f"    {i}. Close {BROWSER_INFO[b]['name']} and use it")
    print(f"    {len(installed) + 1}. Cancel and close a browser yourself first\n")

    while True:
        choice = input(f"  Pick (1-{len(installed) + 1}): ").strip()
        if choice.isdigit():
            idx = int(choice) - 1
            if idx == len(installed):
                return None, False
            if 0 <= idx < len(installed):
                chosen = installed[idx]
                print(f"  → Will close and use {BROWSER_INFO[chosen]['name']}")
                return chosen, True
        print(f"  Please enter a number 1-{len(installed) + 1}")


def show_no_browser_message():
    print(f"\n{'='*60}")
    print(f"  ❌ No supported browser found!")
    print(f"{'='*60}")
    print(f"\n  This script needs one of the following browsers:\n")
    print(f"  • Microsoft Edge   — Usually pre-installed on Windows 10/11")
    print(f"  • Google Chrome    — https://example.com/internal-portal")
    print(f"  • Mozilla Firefox  — https://example.com/internal-portal")
    print(f"  Download: You can also install via Command Prompt (CMD):\n")
    print(f"    How to open CMD:")
    print(f"      1. Press  Win + R  on your keyboard")
    print(f"      2. Type   cmd   and press Enter\n")
    print(f"    Then run one of these commands:")
    print(f"      winget install Mozilla.Firefox")
    print(f"      winget install Google.Chrome")
    print(f"      winget install Microsoft.Edge\n")
    print(f"    Wait for it to finish, then re-run this script.")
    print(f"{'='*60}")


def kill_browser(browser_key):
    info = BROWSER_INFO[browser_key]
    print(f"  [SETUP] Closing {info['name']}...")
    subprocess.run(["taskkill", "/F", "/IM", info["process"]], capture_output=True)
    subprocess.run(["taskkill", "/F", "/IM", info["driver_process"]], capture_output=True)
    time.sleep(5)


def kill_stale_driver(browser_key):
    info = BROWSER_INFO[browser_key]
    subprocess.run(["taskkill", "/F", "/IM", info["driver_process"]], capture_output=True)
    time.sleep(2)


def launch_browser(browser_key, headless=False):
    try:
        if browser_key == "firefox":
            from selenium.webdriver.firefox.options import Options
            options = Options()
            if headless:
                options.add_argument("--headless")
            options.set_preference("dom.webnotifications.enabled", False)
            options.add_argument("--width=1920")
            options.add_argument("--height=1080")
            driver = webdriver.Firefox(options=options)

        elif browser_key == "chrome":
            from selenium.webdriver.chrome.options import Options
            options = Options()
            if headless:
                options.add_argument("--headless=new")
            options.add_argument("--window-size=1920,1080")
            options.add_argument("--disable-notifications")
            options.add_argument("--disable-infobars")
            options.add_argument("--no-first-run")
            options.add_argument("--no-default-browser-check")
            options.add_experimental_option("excludeSwitches", ["enable-logging"])
            driver = webdriver.Chrome(options=options)

        elif browser_key == "edge":
            from selenium.webdriver.edge.options import Options
            options = Options()
            if headless:
                options.add_argument("--headless=new")
            options.add_argument("--window-size=1920,1080")
            options.add_argument("--disable-notifications")
            options.add_argument("--disable-infobars")
            options.add_argument("--no-first-run")
            options.add_argument("--no-default-browser-check")
            options.add_experimental_option("excludeSwitches", ["enable-logging"])
            driver = webdriver.Edge(options=options)

        else:
            print(f"  ❌ Unknown browser: {browser_key}")
            return None

        return driver

    except Exception as e:
        print(f"  ❌ Failed to launch {BROWSER_INFO[browser_key]['name']}: {e}")
        return None


def try_launch_with_fallback(browser_key, headless=False):
    driver = launch_browser(browser_key, headless)
    if driver:
        return driver, browser_key

    print(f"  ⚠️  {BROWSER_INFO[browser_key]['name']} failed, trying alternatives...\n")

    installed = get_installed_browsers()
    running = get_running_browsers()
    alternatives = [b for b in installed if b != browser_key and b not in running]

    for alt in alternatives:
        print(f"  Trying {BROWSER_INFO[alt]['name']}...")
        kill_stale_driver(alt)
        driver = launch_browser(alt, headless)
        if driver:
            print(f"  ✅ Fell back to {BROWSER_INFO[alt]['name']}")
            return driver, alt

    return None, None


# ====================================================================
# CONFIG
# ====================================================================
USERNAME = getpass.getuser()
OUTPUT_DIR = f"C:\\Users\\{USERNAME}\\Documents\\location_data"
COOKIES_DIR = f"C:\\Users\\{USERNAME}\\Documents\\location_data\\cookies"
TIMESTAMP = local_now().strftime("%Y%m%d_%H%M%S")

NA_COLUMNS = [
    ("Location",                                    "Location",                                          "location"),
    ("Enable Core Integration",                         "[Feature] Enable Core Integration",                     "feature"),
    ("Delivery Cluster A",                   "[Cluster] Delivery Cluster A",               "cluster"),
    ("Delivery Cluster B",                "[Cluster] Delivery Cluster B",            "cluster"),
    ("Delivery Cluster C",                     "[Cluster] Delivery Cluster C",                 "cluster"),
    ("Delivery Cluster D",                    "[Cluster] Delivery Cluster D",                "cluster"),
    ("Enable Program A",                        "[Program] Enable Program A",                    "program"),
    ("Enable Program B",                        "[Program] Enable Program B",                    "program"),
    ("Enable Return Program",                    "[Program] Enable Return Program",                "program"),
    ("Enable Unscheduled Return",                  "[Program] Enable Unscheduled Return",              "program"),
    ("Route Type",                                      "[Value] Route Type",                                    "value"),
    ("Enable Recycling Program",                   "[Program] Enable Recycling Program",               "program"),
    ("Enable Service Program A",                      "[Program] Enable Service Program A",                  "program"),
    ("Assign Heavy Items to Cluster A",           "[Program] Assign Heavy Items to Cluster A",      "program"),
    ("Assign Heavy Items to Cluster D",                  "[Program] Assign Heavy Items to Cluster D",              "program"),
    ("Assign Service Items to Cluster D",                  "[Program] Assign Service Items to Cluster D",              "program"),
    ("Assign Recycling to Cluster A",               "[Program] Assign Recycling to Cluster A",           "program"),
    ("Enable Scheduled Delivery",           "[Program] Enable Scheduled Delivery",      "program"),
    ("Enable Length Restriction",    "[Program] Enable Length Restriction", "program"),
    ("Max Weight",                                 "[Value] Max Weight",                               "value"),
    ("Update Partner Logic",                     "[Program] Update Partner Logic",                 "program"),
    ("Single-Driver Max Weight",                              "[Value] Single-Driver Max Weight",                            "value"),
    ("Enable Flexible Capacity",                                "[Program] Enable Flexible Capacity",                            "program"),
    ("Flexible Capacity Max Weight",                            "[Value] Flexible Capacity Max Weight",                          "value"),
]

EU_COLUMNS = [
    ("Location",                                    "Location",                                          "location"),
    ("Enable Core Integration",                         "[Feature] Enable Core Integration",                     "eu"),
    ("Returns Cluster",           "[Cluster] Returns Cluster",     "eu"),
    ("Service Partner Cluster",    "[Cluster] Service Partner Cluster", "eu"),
    ("Service Cluster A",                 "[Cluster] Service Cluster A",             "eu"),
    ("Service Cluster B",                "[Cluster] Service Cluster B",            "eu"),
    ("Two-Person Delivery Cluster",                "[Cluster] Two-Person Delivery Cluster",            "eu"),
    ("One-Person Delivery Cluster",                "[Cluster] One-Person Delivery Cluster",            "eu"),
    ("Enable Returns",                          "[Program] Enable Returns",                      "eu"),
    ("Enable Service Grouping",                "[Program] Enable Service Grouping",             "eu"),
    ("Enable Equipment Option",                 "[Program] Enable Equipment Option",             "eu"),
    ("Enable Regional Clusters",                       "[Program] Enable Regional Clusters",                   "eu"),
    ("Enable Grouping",                      "[Program] Enable Grouping",                  "eu"),
    ("Enable Cross-Border Flow",        "[Program] Enable Cross-Border Flow",    "eu"),
    ("Country",                                    "[Value] Country",                                  "eu_value"),
    ("Enable Service Transfer",                           "[Program] Enable Service Transfer",                       "eu"),
    ("Service Types",                               "[Value] Service Types",                             "eu_value"),
]

NA_FEATURES = ["Enable Core Integration"]
NA_CLUSTERS = ["Delivery Cluster A", "Delivery Cluster B", "Delivery Cluster C", "Delivery Cluster D"]
NA_PROGRAMS = ["Enable Program A", "Enable Program B", "Enable Return Program", "Enable Unscheduled Return", "Enable Recycling Program", "Enable Service Program A", "Assign Heavy Items to Cluster A", "Assign Heavy Items to Cluster D", "Assign Service Items to Cluster D", "Assign Recycling to Cluster A", "Enable Scheduled Delivery", "Enable Length Restriction", "Update Partner Logic", "Enable Flexible Capacity"]
NA_VALUES = ["Route Type", "Max Weight", "Single-Driver Max Weight", "Flexible Capacity Max Weight"]

EU_FEATURES = ["Enable Core Integration"]
EU_CLUSTERS = ["Returns Cluster", "Service Partner Cluster", "Service Cluster A", "Service Cluster B", "Two-Person Delivery Cluster", "One-Person Delivery Cluster"]
EU_PROGRAMS = ["Enable Returns", "Enable Service Grouping", "Enable Equipment Option", "Enable Regional Clusters", "Enable Grouping", "Enable Cross-Border Flow", "Enable Service Transfer"]
EU_VALUES = ["Country", "Service Types"]

REGIONS = {
    "NA": {
        "url": "https://example.com/internal-portal/configuration-review/",
        "cookies_file": "cookies_na.pkl",
        "features": NA_FEATURES, "clusters": NA_CLUSTERS,
        "programs": NA_PROGRAMS, "values": NA_VALUES,
        "columns": NA_COLUMNS,
        "locations": [
           "LOC002", "LOC003", "LOC004", "LOC005", "LOC006", "LOC007",
           "LOC010", "LOC011", "LOC012", "LOC013", "LOC014", "LOC015", "LOC025",
           "LOC026", "LOC027", "LOC028", "LOC029", "LOC030", "LOC032", "LOC033",
           "LOC034", "LOC035", "LOC036", "LOC037", "LOC038", "LOC040",
           "LOC043", "LOC044", "LOC045", "LOC047", "LOC048", "LOC052", "LOC053",
           "LOC057", "LOC058", "LOC059", "LOC060", "LOC061", "LOC062", "LOC066",
           "LOC067", "LOC068", "LOC069", "LOC070", "LOC072", "LOC073", "LOC075",
           "LOC076", "LOC077", "LOC078", "LOC080", "LOC089", "LOC090", "LOC091",
           "LOC092", "LOC093", "LOC094", "LOC095", "LOC097", "LOC098", "LOC099",
           "LOC100", "LOC102", "LOC103", "LOC105", "LOC106", "LOC107", "LOC109",
           "LOC110", "LOC112", "LOC114", "LOC117", "LOC118", "LOC119", "LOC120",
           "LOC121", "LOC122", "LOC126", "LOC129", "LOC132", "LOC135", "LOC139",
           "LOC144", "LOC145", "LOC146", "LOC148", "LOC149", "LOC150",
        ],
    },
    "EU": {
        "url": "https://example.com/internal-portal/configuration-review/",
        "cookies_file": "cookies_eu.pkl",
        "features": EU_FEATURES, "clusters": EU_CLUSTERS,
        "programs": EU_PROGRAMS, "values": EU_VALUES,
        "columns": EU_COLUMNS,
        "locations": [
           "LOC001", "LOC008", "LOC009", "LOC016", "LOC017", "LOC018", "LOC019", "LOC020",
           "LOC021", "LOC022", "LOC023", "LOC024", "LOC031", "LOC039", "LOC041",
           "LOC042", "LOC046", "LOC049", "LOC050", "LOC051", "LOC054", "LOC055",
           "LOC056", "LOC063", "LOC064", "LOC065", "LOC071", "LOC074", "LOC079", "LOC081",
           "LOC082", "LOC083", "LOC084", "LOC085", "LOC086", "LOC087", "LOC088",
           "LOC096", "LOC101", "LOC104", "LOC108", "LOC111", "LOC113", "LOC115",
           "LOC116", "LOC123", "LOC124", "LOC125", "LOC127", "LOC128", "LOC130",
           "LOC131", "LOC133", "LOC134", "LOC136", "LOC137", "LOC138", "LOC140", "LOC141",
           "LOC142", "LOC143", "LOC147",
        ],
    },
}

# ====================================================================
# EMAIL CONFIG
# ====================================================================
EMAIL_RECIPIENTS = [
    "team@example.com", "team@example.com", "team@example.com",
]
EMAIL_CC = [
    "team@example.com",
    "team@example.com",
]
EMAIL_SUBJECT_TEMPLATE = "Configuration Audit Automatic Report — {date}"

OSS_DASHBOARD_URL = (
    "https://example.com/internal-portal"
    "demo-bi-dashboard/workspaces/demo/dashboards/"
    "81812f50-d3f3-4f6b-875b-141eb37f480a"
)
APS_CONFIG_FOLDER_URL = (
    "https://example.com/internal-portal"
    "IgDTD1goga2pS7Kg_MEDfegEAX7fxXnYBsy2nAdvQHW_pRE?e=ZmqMi2"
)

EMAIL_BODY_HTML_TEMPLATE = """\
<html>
<body style="font-family: Calibri, Arial, sans-serif; font-size: 11pt; color: #000000;">
<p>Hi team,</p>

<p>
Attached is the output of the Configuration Audit Automatic Reporting Product.
This information is also viewable in the
<a href="{oss_url}">Operations Dashboard</a>
under NA/EU Configuration Audit tab. A Cloud Drive File folder has been created to store
the excel formatted file. The older link and name being
<a href="{aps_folder_url}">Configuration Audit Data</a>
upon completion of a run. Files are datetime stamped in the file name.
If access to either link is unavailable please contact analyst@example.com (for both)
or dashboard-owner@example.com (for dashboard).
</p>

<p><b>Summary:</b></p>
<ul style="list-style-type: disc; margin-left: 20px;">
  <li>NA: {na_success}/{na_total} locations{na_note}</li>
  <li>EU: {eu_success}/{eu_total} locations{eu_note}</li>
  <li>Generated (Local): {timestamp_local}</li>
  <li>Generated (UTC): {timestamp_utc}</li>
</ul>

{readonly_note}{failed_note}

<p><i>*This is an automatic email sent out by the python product.</i></p>
</body>
</html>
"""

# ====================================================================
# CLOUD_DRIVE UPLOAD CONFIG
# ====================================================================
DOC_PORTAL_HOST = "cloud-documents.example.com"
DOC_PORTAL_PERSONAL_PATH = "/personal/demo_user-retail_com"
CLOUD_DRIVE_FOLDER_NAME = "Configuration Audit Data"
DOC_PORTAL_FOLDER_REL = f"{DOC_PORTAL_PERSONAL_PATH}/Documents/{CLOUD_DRIVE_FOLDER_NAME}"
DOC_PORTAL_API_BASE = f"https://example.com/internal-portal"


def find_local_cloud_drive_folder():
    user_profile = os.environ.get("USERPROFILE", f"C:\\Users\\{USERNAME}")
    roots = []
    for env_key in ("Cloud DriveCommercial", "Cloud Drive", "Cloud DriveConsumer"):
        val = os.environ.get(env_key, "")
        if val and os.path.isdir(val):
            roots.append(val)
    try:
        for entry in os.listdir(user_profile):
            full = os.path.join(user_profile, entry)
            if os.path.isdir(full) and "cloud_drive" in entry.lower():
                roots.append(full)
    except Exception:
        pass
    seen = set()
    unique_roots = []
    for r in roots:
        norm = os.path.normcase(r)
        if norm not in seen:
            seen.add(norm)
            unique_roots.append(r)
    for root in unique_roots:
        candidate = os.path.join(root, CLOUD_DRIVE_FOLDER_NAME)
        if os.path.isdir(candidate):
            return candidate
        candidate = os.path.join(root, "Documents", CLOUD_DRIVE_FOLDER_NAME)
        if os.path.isdir(candidate):
            return candidate
        try:
            for sub in os.listdir(root):
                sub_full = os.path.join(root, sub)
                if os.path.isdir(sub_full):
                    candidate = os.path.join(sub_full, CLOUD_DRIVE_FOLDER_NAME)
                    if os.path.isdir(candidate):
                        return candidate
        except Exception:
            pass
    return None


def try_local_cloud_drive_copy(file_paths):
    folder = find_local_cloud_drive_folder()
    if not folder:
        return False
    print(f"  📂 Found local Cloud Drive sync folder:")
    print(f"     {folder}")
    all_ok = True
    for fp in file_paths:
        if not os.path.exists(fp):
            continue
        dest = os.path.join(folder, os.path.basename(fp))
        try:
            shutil.copy2(fp, dest)
            print(f"  ✅ Copied → {os.path.basename(fp)}")
        except Exception as e:
            print(f"  ❌ Failed to copy {os.path.basename(fp)}: {e}")
            all_ok = False
    if all_ok:
        print(f"  ☁️  Cloud Drive will auto-sync these files to the cloud.\n")
    return all_ok


def extract_document_portal_cookies(driver):
    if not requests_ok:
        return None
    original_url = driver.current_url
    try:
        driver.get(f"https://example.com/internal-portal")
        time.sleep(5)
        if is_csa_blocked(driver):
            driver.get(original_url)
            return None
        current = driver.current_url.lower()
        if "login" in current or "adfs" in current or "microsoftonline" in current:
            time.sleep(15)
            current = driver.current_url.lower()
            if "login" in current or "adfs" in current:
                driver.get(original_url)
                return None
        all_cookies = driver.get_cookies()
        jar = {}
        for c in all_cookies:
            jar[c["name"]] = c["value"]
        if "FedAuth" not in jar and "rtFa" not in jar:
            driver.get(original_url)
            return None
        driver.get(original_url)
        time.sleep(2)
        return jar
    except Exception:
        try:
            driver.get(original_url)
        except Exception:
            pass
        return None


def try_document_portal_api_upload(driver, file_paths):
    if not requests_ok:
        return False
    import requests as req_lib
    print(f"  🔑 Extracting Cloud Document Portal auth from browser session...")
    cookies = extract_document_portal_cookies(driver)
    if not cookies:
        return False
    print(f"  ✅ Got Cloud Document Portal auth cookies")
    digest_url = f"{DOC_PORTAL_API_BASE}/contextinfo"
    headers = {
        "Accept": "application/json;odata=verbose",
        "Content-Type": "application/json;odata=verbose",
    }
    try:
        resp = req_lib.post(digest_url, headers=headers, cookies=cookies, timeout=30)
        if resp.status_code != 200:
            return False
        digest = resp.json()["d"]["GetContextWebInformation"]["FormDigestValue"]
    except Exception:
        return False
    folder_check_url = (
        f"{DOC_PORTAL_API_BASE}/web/GetFolderByServerRelativeUrl"
        f"('{DOC_PORTAL_FOLDER_REL}')"
    )
    try:
        resp = req_lib.get(folder_check_url, headers=headers, cookies=cookies, timeout=30)
        if resp.status_code == 404:
            create_url = (
                f"{DOC_PORTAL_API_BASE}/web/GetFolderByServerRelativeUrl"
                f"('{DOC_PORTAL_PERSONAL_PATH}/Documents')/folders/add"
                f"('{CLOUD_DRIVE_FOLDER_NAME}')"
            )
            create_headers = {**headers, "X-RequestDigest": digest}
            resp2 = req_lib.post(create_url, headers=create_headers, cookies=cookies, timeout=30)
            if resp2.status_code not in (200, 201):
                return False
    except Exception:
        return False
    any_success = False
    for fp in file_paths:
        if not os.path.exists(fp):
            continue
        filename = os.path.basename(fp)
        file_size = os.path.getsize(fp)
        upload_url = (
            f"{DOC_PORTAL_API_BASE}/web/GetFolderByServerRelativeUrl"
            f"('{DOC_PORTAL_FOLDER_REL}')/Files/add"
            f"(url='{filename}',overwrite=true)"
        )
        upload_headers = {
            "Accept": "application/json;odata=verbose",
            "X-RequestDigest": digest,
            "Content-Length": str(file_size),
        }
        try:
            with open(fp, "rb") as f:
                file_bytes = f.read()
            print(f"  ⬆️  Uploading {filename} ({file_size:,} bytes)...")
            resp = req_lib.post(upload_url, headers=upload_headers, cookies=cookies, data=file_bytes, timeout=120)
            if resp.status_code in (200, 201):
                print(f"  ✅ Uploaded → {filename}")
                any_success = True
            else:
                print(f"  ❌ Upload failed (HTTP {resp.status_code})")
        except Exception as e:
            print(f"  ❌ Upload error for {filename}: {e}")
    return any_success


def try_browser_upload_fallback(file_paths):
    print(f"\n  📂 Opening Cloud Drive folder in your browser for manual upload...")
    print(f"  📂 Opening local file location in Explorer...")
    print()
    print(f"  Steps:")
    print(f"    1. A browser tab will open to the Cloud Drive folder")
    print(f"    2. An Explorer window will open to your local files")
    print(f"    3. Drag the files from Explorer into the Cloud Drive browser tab")
    print(f"    4. Wait for the upload to complete\n")
    try:
        os.startfile(APS_CONFIG_FOLDER_URL)
    except Exception:
        print(f"     Open manually: {APS_CONFIG_FOLDER_URL}")
    time.sleep(2)
    if file_paths:
        local_dir = os.path.dirname(file_paths[0])
        try:
            subprocess.Popen(["explorer", local_dir])
        except Exception:
            print(f"     Files are at: {local_dir}")
    return True


def upload_to_cloud_drive(driver, file_paths):
    print(f"\n{'='*60}")
    print(f"  ☁️  UPLOADING TO CLOUD_DRIVE")
    print(f"{'='*60}\n")
    existing = [fp for fp in file_paths if os.path.exists(fp)]
    if not existing:
        print(f"  ⚠️  No files found to upload.")
        return "failed"
    for fp in existing:
        print(f"  📄 {os.path.basename(fp)}")
    print()
    print(f"  🔍 Method 1: Checking for local Cloud Drive sync folder...")
    try:
        if try_local_cloud_drive_copy(existing):
            return "synced"
        else:
            print(f"     ℹ️  Local sync folder not found — trying API upload.\n")
    except Exception as e:
        print(f"     ⚠️  Local copy error: {e}\n")
    print(f"  🔍 Method 2: Trying Cloud Document Portal API upload...")
    try:
        if driver is not None and try_document_portal_api_upload(driver, existing):
            print(f"  ✅ Files uploaded via Cloud Document Portal API!\n")
            return "api"
        else:
            print(f"     ℹ️  API upload unsuccessful — falling back to manual.\n")
    except Exception as e:
        print(f"     ⚠️  API upload error: {e}\n")
    print(f"  🔍 Method 3: Manual upload (drag & drop)...")
    try_browser_upload_fallback(existing)
    return "manual"


# ====================================================================
# CORE FUNCTIONS
# ====================================================================
def is_on_actual_site(url, region):
    base = REGIONS[region]["url"].replace("/configuration-review/", "")
    return url.startswith(base)


def save_cookies(driver, region):
    os.makedirs(COOKIES_DIR, exist_ok=True)
    path = os.path.join(COOKIES_DIR, REGIONS[region]["cookies_file"])
    with open(path, "wb") as f:
        pickle.dump(driver.get_cookies(), f)


def load_cookies(driver, region):
    path = os.path.join(COOKIES_DIR, REGIONS[region]["cookies_file"])
    if not os.path.exists(path):
        return False
    if time.time() - os.path.getmtime(path) > 43200:
        return False
    try:
        with open(path, "rb") as f:
            cookies = pickle.load(f)
        base = REGIONS[region]["url"].replace("/configuration-review/", "")
        driver.get(base)
        time.sleep(2)
        for c in cookies:
            try:
                driver.add_cookie(c)
            except Exception:
                pass
        return True
    except Exception:
        return False


def login_with_browser(driver, region, browser_name):
    print()
    print(f"  {'='*56}")
    print(f"  🔐  LOGIN REQUIRED FOR {region}")
    print(f"  {'='*56}")
    print()
    print(f"  A {browser_name} window should now be open.")
    print(f"  If you don't see it, check your taskbar at the bottom")
    print(f"  of your screen — it may be behind other windows.")
    print()
    print(f"  Steps:")
    print(f"    1. Click on the {browser_name} window")
    print(f"    2. Enter your username and password")
    print(f"    3. Complete MFA (tap security key, approve push, etc.)")
    print(f"    4. Come back here — the script detects login automatically")
    print()
    print(f"  ⏳ Waiting up to 5 minutes...")
    print(f"  💡 You do NOT need to watch the browser after logging in.")
    print(f"     Once you log in, feel free to minimize it and keep")
    print(f"     working on other things. The script handles the rest.")
    print()

    try:
        driver.maximize_window()
        driver.switch_to.window(driver.current_window_handle)
    except Exception:
        pass

    try:
        import winsound
        winsound.MessageBeep(winsound.MB_ICONEXCLAMATION)
    except Exception:
        pass

    start = time.time()
    last_print = 0
    while time.time() - start < 300:
        if is_csa_blocked(driver):
            print(f"\n  🔒 CSA block page detected — VPN may have dropped!")
            vpn_ok = prompt_vpn_connection_browser(driver, REGIONS[region]["url"])
            if not vpn_ok:
                return False
            if is_on_actual_site(driver.current_url, region):
                print(f"  ✅ Logged in to {region} successfully!")
                print(f"  💡 The browser will now work on its own. Feel free to")
                print(f"     minimize it and continue with your other work.\n")
                time.sleep(10)
                save_cookies(driver, region)
                return True
            print(f"  ✅ VPN connected! Now waiting for login...")
            print(f"  💡 Complete your login in the {browser_name} window.\n")
            start = time.time()
            last_print = 0
            continue

        if is_on_actual_site(driver.current_url, region):
            print(f"  ✅ Logged in to {region} successfully!")
            print(f"  💡 The browser will now work on its own. Feel free to")
            print(f"     minimize it and continue with your other work.")
            print(f"     The script will notify you when it's done.\n")

            try:
                import winsound
                winsound.MessageBeep(winsound.MB_OK)
            except Exception:
                pass

            time.sleep(10)
            save_cookies(driver, region)
            return True

        elapsed = int(time.time() - start)
        if elapsed - last_print >= 15:
            remaining = 300 - elapsed
            mins = remaining // 60
            secs = remaining % 60
            print(f"  ⏳ Still waiting for login... {mins}m {secs}s remaining")
            last_print = elapsed

        time.sleep(2)

    print(f"\n  ❌ Login timed out after 5 minutes.")
    print(f"  Make sure you completed login in the {browser_name} window.")
    return False


# ====================================================================
# EDIT BUTTON DETECTION
# ====================================================================
def try_click_edit(driver):
    for xpath in [
        "//button[contains(text(),'Edit Location')]",
        "//a[contains(text(),'Edit Location')]",
        "//*[contains(text(),'Edit Location')]",
    ]:
        try:
            el = driver.find_element(By.XPATH, xpath)
            el.click()
            return "edit"
        except NoSuchElementException:
            continue

    for btn in driver.find_elements(By.TAG_NAME, "button"):
        txt = btn.text.strip().lower()
        if "edit" in txt and "location" in txt:
            btn.click()
            return "edit"

    for xpath in [
        "//button[contains(text(),'View Location')]",
        "//a[contains(text(),'View Location')]",
        "//*[contains(text(),'View Location')]",
    ]:
        try:
            el = driver.find_element(By.XPATH, xpath)
            el.click()
            return "view"
        except NoSuchElementException:
            continue

    for btn in driver.find_elements(By.TAG_NAME, "button"):
        txt = btn.text.strip().lower()
        if "view" in txt and "location" in txt:
            btn.click()
            return "view"

    return None


def check_has_edit_permission(driver):
    page_text = driver.find_element(By.TAG_NAME, "body").text.lower()
    if "edit location" in page_text:
        return "edit"
    if "view location" in page_text:
        return "view"
    return None


# ====================================================================
# READ-ONLY SCRAPE
# ====================================================================
def scrape_location_readonly(driver, wait, location, region):
    config = REGIONS[region]
    row = {"Location": location}

    for f in config["features"]:
        row[f"[Feature] {f}"] = "No Edit Access"
    for c in config["clusters"]:
        row[f"[Cluster] {c}"] = "No Edit Access"
    for p in config["programs"]:
        row[f"[Program] {p}"] = "No Edit Access"
    for v in config["values"]:
        row[f"[Value] {v}"] = ""

    location_input = driver.find_element(By.CSS_SELECTOR, "input[placeholder='Location Code']")
    location_input.click()
    time.sleep(0.5)
    location_input.send_keys(Keys.CONTROL + "a")
    location_input.send_keys(Keys.DELETE)
    time.sleep(0.5)
    location_input.send_keys(location)
    time.sleep(1)

    try:
        option = wait.until(EC.element_to_be_clickable(
            (By.XPATH, f"//*[contains(text(),'{location}')]")
        ))
        option.click()
    except TimeoutException:
        raise Exception(f"'{location}' not found in dropdown")

    time.sleep(3)

    if is_csa_blocked(driver):
        raise VPNRequiredError("VPN disconnected — CSA page detected")

    view_clicked = False
    for xpath in [
        "//button[contains(text(),'View Location')]",
        "//a[contains(text(),'View Location')]",
        "//*[contains(text(),'View Location')]",
    ]:
        try:
            el = driver.find_element(By.XPATH, xpath)
            el.click()
            view_clicked = True
            break
        except NoSuchElementException:
            continue

    if not view_clicked:
        for btn in driver.find_elements(By.TAG_NAME, "button"):
            txt = btn.text.strip().lower()
            if "view" in txt and "location" in txt:
                btn.click()
                view_clicked = True
                break

    if view_clicked:
        time.sleep(4)

    page_text = driver.find_element(By.TAG_NAME, "body").text
    lines = page_text.split("\n")

    toggles = driver.find_elements(By.CSS_SELECTOR, "[role='switch']")
    for toggle in toggles:
        label = ""
        try:
            parent = toggle.find_element(By.XPATH, "./..")
            label = parent.text.strip()
        except Exception:
            pass
        if not label:
            try:
                gp = toggle.find_element(By.XPATH, "./../..")
                label = gp.text.strip()
            except Exception:
                pass
        if not label:
            continue

        checked = toggle.get_attribute("aria-checked") == "true"
        status = "Active" if checked else "Inactive"

        for f in config["features"]:
            if f.lower() in label.lower():
                row[f"[Feature] {f}"] = status
                break
        for p in config["programs"]:
            if p.lower() in label.lower():
                row[f"[Program] {p}"] = status
                break

    checkboxes = driver.find_elements(By.CSS_SELECTOR, "input[type='checkbox']")
    for cb in checkboxes:
        label = ""
        try:
            parent = cb.find_element(By.XPATH, "./..")
            label = parent.text.strip()
        except Exception:
            pass
        if not label:
            try:
                gp = cb.find_element(By.XPATH, "./../..")
                label = gp.text.strip()
            except Exception:
                pass
        if not label:
            continue

        checked = cb.get_attribute("aria-checked") == "true" or cb.is_selected()
        status = "Active" if checked else "Inactive"

        for c in config["clusters"]:
            if c.lower() in label.lower() or label.lower() in c.lower():
                row[f"[Cluster] {c}"] = status
                break

    if region == "NA":
        for i, line in enumerate(lines):
            stripped = line.strip()
            if stripped == "Route Type" and i + 1 < len(lines):
                row["[Value] Route Type"] = lines[i + 1].strip()
            if stripped == "Max Weight" and "1P" not in lines[max(i - 1, 0)].strip():
                for j in range(i + 1, min(i + 3, len(lines))):
                    val = lines[j].strip()
                    if val.replace(".", "").isdigit():
                        row["[Value] Max Weight"] = f"{val} lb"
                        break
            if stripped == "Single-Driver Max Weight":
                for j in range(i + 1, min(i + 3, len(lines))):
                    val = lines[j].strip()
                    if val.replace(".", "").isdigit():
                        row["[Value] Single-Driver Max Weight"] = f"{val} lb"
                        break
            if stripped == "Flexible Capacity Max Weight":
                for j in range(i + 1, min(i + 3, len(lines))):
                    val = lines[j].strip()
                    if val.replace(".", "").isdigit():
                        row["[Value] Flexible Capacity Max Weight"] = f"{val} lb"
                        break

    if region == "EU":
        for i, line in enumerate(lines):
            stripped = line.strip()
            if stripped == "Country" and i + 1 < len(lines):
                row["[Value] Country"] = lines[i + 1].strip()
            if stripped == "Service Types" and i + 1 < len(lines):
                services = []
                for j in range(i + 1, min(i + 10, len(lines))):
                    svc = lines[j].strip()
                    if svc in ["Cancel", "Review Changes", "Close", ""]:
                        break
                    if "selected" not in svc:
                        services.append(svc)
                if services:
                    row["[Value] Service Types"] = ", ".join(services)

    page_lower = page_text.lower()
    all_items = (
        [("feature", f) for f in config["features"]]
        + [("program", p) for p in config["programs"]]
        + [("cluster", c) for c in config["clusters"]]
    )
    for item_type, item_name in all_items:
        if item_type == "feature":
            key = f"[Feature] {item_name}"
        elif item_type == "program":
            key = f"[Program] {item_name}"
        else:
            key = f"[Cluster] {item_name}"

        if row[key] != "No Edit Access":
            continue

        if item_name.lower() in page_lower:
            idx = page_lower.find(item_name.lower())
            surrounding = page_lower[max(0, idx - 50):idx + len(item_name) + 50]
            if "enabled" in surrounding or "active" in surrounding or "true" in surrounding:
                row[key] = "Active"
            elif "disabled" in surrounding or "inactive" in surrounding or "false" in surrounding:
                row[key] = "Inactive"
            else:
                row[key] = "Visible (state unknown)"

    closed = False
    for btn in driver.find_elements(By.TAG_NAME, "button"):
        txt = btn.text.strip().lower()
        if txt in ["close", "cancel", "back"]:
            btn.click()
            closed = True
            break
    if not closed:
        driver.get(config["url"])

    time.sleep(2)
    return row


# ====================================================================
# FULL EDIT-MODE SCRAPE
# ====================================================================
def scrape_location(driver, wait, location, region):
    config = REGIONS[region]
    row = {"Location": location}
    for f in config["features"]:
        row[f"[Feature] {f}"] = "Unavailable"
    for c in config["clusters"]:
        row[f"[Cluster] {c}"] = "Unavailable"
    for p in config["programs"]:
        row[f"[Program] {p}"] = "Unavailable"
    for v in config["values"]:
        row[f"[Value] {v}"] = ""

    location_input = driver.find_element(By.CSS_SELECTOR, "input[placeholder='Location Code']")
    location_input.click()
    time.sleep(0.5)
    location_input.send_keys(Keys.CONTROL + "a")
    location_input.send_keys(Keys.DELETE)
    time.sleep(0.5)
    location_input.send_keys(location)
    time.sleep(1)

    try:
        option = wait.until(EC.element_to_be_clickable(
            (By.XPATH, f"//*[contains(text(),'{location}')]")
        ))
        option.click()
    except TimeoutException:
        raise Exception(f"'{location}' not found in dropdown")

    time.sleep(3)

    if is_csa_blocked(driver):
        raise VPNRequiredError("VPN disconnected — CSA page detected")

    result = try_click_edit(driver)

    if result is None:
        raise NoEditPermissionError(
            "No Edit or View button found — you may lack permissions"
        )
    if result == "view":
        raise NoEditPermissionError(
            "Only 'View Location' available — no edit permissions"
        )

    time.sleep(4)

    if is_csa_blocked(driver):
        raise VPNRequiredError("VPN disconnected — CSA page detected after clicking Edit")

    toggles = driver.find_elements(By.CSS_SELECTOR, "[role='switch']")
    for toggle in toggles:
        label = ""
        try:
            parent = toggle.find_element(By.XPATH, "./..")
            label = parent.text.strip()
        except Exception:
            pass
        if not label:
            try:
                gp = toggle.find_element(By.XPATH, "./../..")
                label = gp.text.strip()
            except Exception:
                pass
        if not label:
            continue

        checked = toggle.get_attribute("aria-checked") == "true"
        status = "Active" if checked else "Inactive"

        for f in config["features"]:
            if f.lower() in label.lower():
                row[f"[Feature] {f}"] = status
                break
        for p in config["programs"]:
            if p.lower() in label.lower():
                row[f"[Program] {p}"] = status
                break

    checkboxes = driver.find_elements(By.CSS_SELECTOR, "input[type='checkbox']")
    for cb in checkboxes:
        label = ""
        try:
            parent = cb.find_element(By.XPATH, "./..")
            label = parent.text.strip()
        except Exception:
            pass
        if not label:
            try:
                gp = cb.find_element(By.XPATH, "./../..")
                label = gp.text.strip()
            except Exception:
                pass
        if not label:
            continue

        checked = cb.get_attribute("aria-checked") == "true" or cb.is_selected()
        status = "Active" if checked else "Inactive"

        for c in config["clusters"]:
            if c.lower() in label.lower() or label.lower() in c.lower():
                row[f"[Cluster] {c}"] = status
                break

    page_text = driver.find_element(By.TAG_NAME, "body").text
    lines = page_text.split("\n")

    if region == "NA":
        for i, line in enumerate(lines):
            stripped = line.strip()
            if stripped == "Route Type" and i + 1 < len(lines):
                row["[Value] Route Type"] = lines[i + 1].strip()
            if stripped == "Max Weight" and "1P" not in lines[max(i - 1, 0)].strip():
                for j in range(i + 1, min(i + 3, len(lines))):
                    val = lines[j].strip()
                    if val.replace(".", "").isdigit():
                        row["[Value] Max Weight"] = f"{val} lb"
                        break
            if stripped == "Single-Driver Max Weight":
                for j in range(i + 1, min(i + 3, len(lines))):
                    val = lines[j].strip()
                    if val.replace(".", "").isdigit():
                        row["[Value] Single-Driver Max Weight"] = f"{val} lb"
                        break
            if stripped == "Flexible Capacity Max Weight":
                for j in range(i + 1, min(i + 3, len(lines))):
                    val = lines[j].strip()
                    if val.replace(".", "").isdigit():
                        row["[Value] Flexible Capacity Max Weight"] = f"{val} lb"
                        break

    if region == "EU":
        for i, line in enumerate(lines):
            stripped = line.strip()
            if stripped == "Country" and i + 1 < len(lines):
                row["[Value] Country"] = lines[i + 1].strip()
            if stripped == "Service Types" and i + 1 < len(lines):
                services = []
                for j in range(i + 1, min(i + 10, len(lines))):
                    svc = lines[j].strip()
                    if svc in ["Cancel", "Review Changes", ""]:
                        break
                    if "selected" not in svc:
                        services.append(svc)
                if services:
                    row["[Value] Service Types"] = ", ".join(services)

    cancel_clicked = False
    for btn in driver.find_elements(By.TAG_NAME, "button"):
        if btn.text.strip().lower() == "cancel":
            btn.click()
            cancel_clicked = True
            break
    if not cancel_clicked:
        for xpath in ["//button[contains(text(),'Cancel')]", "//*[contains(text(),'Cancel')]"]:
            try:
                driver.find_element(By.XPATH, xpath).click()
                cancel_clicked = True
                break
            except NoSuchElementException:
                continue
    if not cancel_clicked:
        driver.get(config["url"])

    time.sleep(2)
    return row


# ====================================================================
# EXCEL OUTPUT
# ====================================================================
def create_excel(na_data, eu_data, na_failed, eu_failed, scrape_mode, excel_path):
    wb = Workbook()

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")
    data_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", size=10, bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    na_stripe = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    eu_stripe = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    na_hdr = {
        "location": PatternFill(start_color="333333", end_color="333333", fill_type="solid"),
        "feature": PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid"),
        "cluster": PatternFill(start_color="385723", end_color="385723", fill_type="solid"),
        "program": PatternFill(start_color="7B2D8E", end_color="7B2D8E", fill_type="solid"),
        "value":   PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid"),
    }
    eu_hdr = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    eu_location_hdr = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

    active_font = Font(name="Calibri", size=10, bold=True, color="006100")
    active_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    inactive_font = Font(name="Calibri", size=10, bold=True, color="9C0006")
    inactive_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    unavail_font = Font(name="Calibri", size=10, color="808080", italic=True)
    unavail_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    noedit_font = Font(name="Calibri", size=10, color="C05000", italic=True)
    noedit_fill = PatternFill(start_color="FFE0CC", end_color="FFE0CC", fill_type="solid")
    unknown_font = Font(name="Calibri", size=10, color="4472C4", italic=True)
    unknown_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    value_font = Font(name="Calibri", size=10, bold=True, color="7F6000")
    value_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fail_hdr = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

    def style_cell(c, v, ct, rf):
        c.alignment = center_align
        c.border = thin_border
        if ct == "location":
            c.font, c.fill = bold_font, rf
        elif v == "Active":
            c.font, c.fill = active_font, active_fill
        elif v == "Inactive":
            c.font, c.fill = inactive_font, inactive_fill
        elif v == "Unavailable":
            c.font, c.fill = unavail_font, unavail_fill
        elif v == "No Edit Access":
            c.font, c.fill = noedit_font, noedit_fill
        elif v and "state unknown" in str(v).lower():
            c.font, c.fill = unknown_font, unknown_fill
        elif ct in ("value", "eu_value") and v:
            c.font, c.fill = value_font, value_fill
        else:
            c.font, c.fill = data_font, rf

    def write_na(ws, data):
        for ci, (dn, dk, ct) in enumerate(NA_COLUMNS, 1):
            c = ws.cell(row=1, column=ci, value=dn)
            c.font = header_font
            c.fill = na_hdr.get(ct, na_hdr["program"])
            c.alignment = header_align
            c.border = thin_border
        for ri, rd in enumerate(data, 2):
            rf = na_stripe if ri % 2 == 0 else white
            for ci, (dn, dk, ct) in enumerate(NA_COLUMNS, 1):
                v = rd.get(dk, "")
                c = ws.cell(row=ri, column=ci, value=v)
                style_cell(c, v, ct, rf)
        for ci, (dn, _, ct) in enumerate(NA_COLUMNS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = (
                12 if ct == "location"
                else (16 if ct == "value" else min(max(len(dn) + 4, 18), 35))
            )
        ws.freeze_panes = "B2"
        if data:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(NA_COLUMNS))}{len(data) + 1}"

    def write_eu(ws, data):
        for ci, (dn, dk, ct) in enumerate(EU_COLUMNS, 1):
            c = ws.cell(row=1, column=ci, value=dn)
            c.font, c.alignment, c.border = header_font, header_align, thin_border
            c.fill = eu_location_hdr if ct == "location" else eu_hdr
        for ri, rd in enumerate(data, 2):
            rf = eu_stripe if ri % 2 == 0 else white
            for ci, (dn, dk, ct) in enumerate(EU_COLUMNS, 1):
                v = rd.get(dk, "")
                c = ws.cell(row=ri, column=ci, value=v)
                style_cell(c, v, ct, rf)
        for ci, (dn, _, ct) in enumerate(EU_COLUMNS, 1):
            ws.column_dimensions[get_column_letter(ci)].width = (
                12 if ct == "location"
                else (18 if ct == "eu_value" else min(max(len(dn) + 4, 18), 38))
            )
        ws.freeze_panes = "B2"
        if data:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(EU_COLUMNS))}{len(data) + 1}"

    ws_na = wb.active
    ws_na.title = "NA Locations"
    write_na(ws_na, na_data)

    ws_eu = wb.create_sheet("EU Locations")
    write_eu(ws_eu, eu_data)

    all_failed = (
        [{"Region": "NA", **f} for f in na_failed]
        + [{"Region": "EU", **f} for f in eu_failed]
    )
    if all_failed:
        ws_f = wb.create_sheet("Failed Locations")
        for col, h in enumerate(["Region", "Location", "Error"], 1):
            c = ws_f.cell(row=1, column=col, value=h)
            c.font, c.fill, c.alignment, c.border = header_font, fail_hdr, header_align, thin_border
        for r, fail in enumerate(all_failed, 2):
            for col, key in enumerate(["Region", "location", "error"], 1):
                c = ws_f.cell(row=r, column=col, value=fail.get(key, ""))
                c.border, c.font = thin_border, data_font
        ws_f.column_dimensions["A"].width = 10
        ws_f.column_dimensions["B"].width = 12
        ws_f.column_dimensions["C"].width = 50

    if scrape_mode == "readonly":
        ws_l = wb.create_sheet("⚠️ Read-Only Notice")
        ws_l.cell(row=1, column=1, value="READ-ONLY MODE").font = Font(size=14, bold=True, color="C05000")
        ws_l.cell(row=3, column=1, value="This report was generated without Edit Location permissions.").font = data_font
        ws_l.cell(row=4, column=1, value="Some data may be incomplete or show 'No Edit Access'.").font = data_font
        ws_l.cell(row=5, column=1, value="Request edit permissions from your manager for full data.").font = data_font
        ws_l.cell(row=7, column=1, value="Cell Legend:").font = Font(size=11, bold=True)

        legend = [
            ("Active", active_font, active_fill, "Feature/program is turned ON"),
            ("Inactive", inactive_font, inactive_fill, "Feature/program is turned OFF"),
            ("Unavailable", unavail_font, unavail_fill, "Not found on the page"),
            ("No Edit Access", noedit_font, noedit_fill, "Could not read — no edit permissions"),
            ("Visible (state unknown)", unknown_font, unknown_fill, "Found on page but state unclear"),
        ]
        for i, (label, lfont, lfill, desc) in enumerate(legend, 9):
            c = ws_l.cell(row=i, column=1, value=label)
            c.font, c.fill, c.border = lfont, lfill, thin_border
            ws_l.cell(row=i, column=2, value=desc).font = data_font

        ws_l.column_dimensions["A"].width = 25
        ws_l.column_dimensions["B"].width = 50

    wb.save(excel_path)


# ====================================================================
# OUTLOOK EMAIL DRAFT
# ====================================================================
def draft_outlook_email(excel_path, na_data, eu_data, na_failed, eu_failed,
                        na_mode, eu_mode, overall_mode):
    if not win32_ok:
        print(f"  ⚠️  pywin32 not installed — skipping email draft.")
        print(f"     Install with:  pip install pywin32")
        return False

    try:
        import win32com.client

        na_total = len(REGIONS["NA"]["locations"])
        eu_total = len(REGIONS["EU"]["locations"])
        na_success = na_total - len(na_failed)
        eu_success = eu_total - len(eu_failed)

        na_note = ""
        if na_mode == "readonly":
            na_note = " (read-only)"
        elif na_mode in ("skip", "vpn_failed"):
            na_note = " (skipped)"

        eu_note = ""
        if eu_mode == "readonly":
            eu_note = " (read-only)"
        elif eu_mode in ("skip", "vpn_failed"):
            eu_note = " (skipped)"

        readonly_note = ""
        if overall_mode == "readonly":
            readonly_note = (
                "<p><b>Note:</b> Some data was collected in read-only mode. "
                "Cells marked 'No Edit Access' could not be read.</p>\n"
            )

        total_failed = len(na_failed) + len(eu_failed)
        failed_note = ""
        if total_failed > 0:
            failed_note = (
                f"<p>⚠️ {total_failed} location(s) failed — "
                f"see the 'Failed Locations' tab in the Excel for details.</p>\n"
            )

        subject = EMAIL_SUBJECT_TEMPLATE.format(
            date=local_now().strftime("%Y-%m-%d"),
        )
        html_body = EMAIL_BODY_HTML_TEMPLATE.format(
            oss_url=OSS_DASHBOARD_URL,
            aps_folder_url=APS_CONFIG_FOLDER_URL,
            na_success=na_success,
            na_total=na_total,
            na_note=na_note,
            eu_success=eu_success,
            eu_total=eu_total,
            eu_note=eu_note,
            timestamp_local=fmt_local(),
            timestamp_utc=fmt_utc(),
            readonly_note=readonly_note,
            failed_note=failed_note,
        )

        outlook = win32com.client.Dispatch("Email Client.Application")
        mail = outlook.CreateItem(0)

        mail.To = "; ".join(EMAIL_RECIPIENTS)
        if EMAIL_CC:
            mail.CC = "; ".join(EMAIL_CC)
        mail.Subject = subject
        mail.HTMLBody = html_body

        if os.path.exists(excel_path):
            mail.Attachments.Add(os.path.abspath(excel_path))
        else:
            print(f"  ⚠️  Excel file not found at {excel_path}")
            return False

        mail.Display()

        print(f"  ✅ Email Client draft created!")
        print(f"     To: {', '.join(EMAIL_RECIPIENTS)}")
        if EMAIL_CC:
            print(f"     CC: {', '.join(EMAIL_CC)}")
        print(f"     📎 {os.path.basename(excel_path)} attached")
        print(f"     → Review the email and click Send when ready.\n")
        return True

    except FileNotFoundError:
        print(f"  ⚠️  Email Client does not appear to be installed.")
        return False
    except Exception as e:
        error_str = str(e).lower()
        if "outlook" in error_str or "class not registered" in error_str:
            print(f"  ⚠️  Could not connect to Email Client: {e}")
            print(f"     Make sure Email Client is installed and has been opened at least once.")
        else:
            print(f"  ⚠️  Email draft failed: {e}")
        return False


# ====================================================================
# PERMISSION PROMPTS
# ====================================================================
def show_no_edit_warning(region):
    print(f"\n  {'='*56}")
    print(f"  ⚠️  NO EDIT PERMISSION DETECTED FOR {region}")
    print(f"  {'='*56}")
    print()
    print(f"  The first location did not have an 'Edit Location' button.")
    print(f"  This usually means your account doesn't have edit access")
    print(f"  to the Cluster Transfer tool.\n")
    print(f"  Without edit access, the scraper can only read limited")
    print(f"  data visible on the location's summary/view page.\n")
    print(f"  Options:")
    print(f"    1. Continue in READ-ONLY mode (partial data)")
    print(f"       → Scrapes whatever is visible without Edit")
    print(f"       → Cells will show 'No Edit Access' where data is missing\n")
    print(f"    2. Skip {region} entirely")
    print(f"       → Move on to the next region\n")
    print(f"    3. Stop the program")
    print(f"       → Exit and request edit permissions from your manager\n")

    while True:
        choice = input(f"  Pick (1/2/3): ").strip()
        if choice == "1":
            print(f"\n  → Continuing in read-only mode for {region}...")
            print(f"  💡 This will run automatically now. Feel free to minimize")
            print(f"     this window and keep working on other things.\n")
            return "readonly"
        elif choice == "2":
            print(f"\n  → Skipping {region}...")
            return "skip"
        elif choice == "3":
            print(f"\n  → Stopping. Request edit access and try again.")
            return "abort"
        print(f"  Please enter 1, 2, or 3")


def navigate_back_safely(driver, url):
    try:
        for btn in driver.find_elements(By.TAG_NAME, "button"):
            if btn.text.strip().lower() in ["close", "cancel", "back"]:
                btn.click()
                time.sleep(2)
                return
        driver.get(url)
        time.sleep(5)
    except Exception:
        driver.get(url)
        time.sleep(5)


# ====================================================================
# HEADLESS-TO-VISIBLE RELAUNCH
# ====================================================================
def relaunch_visible(driver, browser_key):
    print(f"\n  🔄 Cookies didn't work — need to log in.")
    print(f"     Restarting browser in visible mode...\n")

    try:
        driver.quit()
    except Exception:
        pass

    time.sleep(3)
    kill_stale_driver(browser_key)

    driver, actual_key = try_launch_with_fallback(browser_key, headless=False)

    if driver:
        actual_name = BROWSER_INFO[actual_key]["name"]
        print(f"  ✅ {actual_name} relaunched in visible mode")
        print(f"  💡 After logging in, you can minimize the browser and")
        print(f"     keep working. The script handles everything else.\n")
    else:
        print(f"  ❌ Could not relaunch browser in visible mode.")

    return driver, actual_key


# ====================================================================
# REGION SCRAPER
# ====================================================================
def scrape_region(driver, wait, region, browser_name, is_headless):
    config = REGIONS[region]
    url, locations = config["url"], config["locations"]
    total = len(locations)
    all_data, failed = [], []
    start_time = time.time()
    scrape_mode = "edit"
    permission_checked = False
    consecutive_no_edit = 0

    print(f"\n{'='*60}")
    print(f"  🌍 {region}: {total} locations")
    print(f"{'='*60}\n")

    driver.get(url)
    time.sleep(10)

    if is_csa_blocked(driver):
        if is_headless:
            print(f"  🔒 CSA block detected in headless mode.")
            return all_data, failed, 0, "needs_relaunch", True

        print(f"  🔒 CSA block page detected — VPN may have dropped!")
        vpn_ok = prompt_vpn_connection_browser(driver, url)
        if not vpn_ok:
            for s in locations:
                failed.append({"location": s, "error": "VPN not connected"})
            return all_data, failed, 0, "vpn_failed", False

    if not is_on_actual_site(driver.current_url, region):
        if is_headless:
            print(f"  🔐 Login required but browser is running in headless mode.")
            print(f"     Cookies may be expired or invalid.")
            return all_data, failed, 0, "needs_relaunch", True

        if not login_with_browser(driver, region, browser_name):
            for s in locations:
                failed.append({"location": s, "error": "Login failed"})
            return all_data, failed, 0, "skip", False
    else:
        print(f"  ✅ {region} site loaded!")
        if is_headless:
            print(f"  💡 Running in the background — no interaction needed.\n")
        else:
            print(f"  💡 The browser is working on its own now. You can minimize")
            print(f"     it and keep working — the script will notify you when done.\n")
        save_cookies(driver, region)

    time.sleep(5)

    print(f"  🔍 Checking edit permissions on first location ({locations[0]})...\n")

    try:
        location_input = driver.find_element(
            By.CSS_SELECTOR, "input[placeholder='Location Code']"
        )
        location_input.click()
        time.sleep(0.5)
        location_input.send_keys(Keys.CONTROL + "a")
        location_input.send_keys(Keys.DELETE)
        time.sleep(0.5)
        location_input.send_keys(locations[0])
        time.sleep(1)

        try:
            option = wait.until(EC.element_to_be_clickable(
                (By.XPATH, f"//*[contains(text(),'{locations[0]}')]")
            ))
            option.click()
        except TimeoutException:
            pass

        time.sleep(3)

        if is_csa_blocked(driver):
            if is_headless:
                return all_data, failed, 0, "needs_relaunch", True

            print(f"  🔒 CSA page appeared — VPN may have disconnected!")
            vpn_ok = prompt_vpn_connection_browser(driver, url)
            if not vpn_ok:
                for s in locations:
                    failed.append({"location": s, "error": "VPN not connected"})
                return all_data, failed, 0, "vpn_failed", False
            driver.get(url)
            time.sleep(5)
            location_input = driver.find_element(
                By.CSS_SELECTOR, "input[placeholder='Location Code']"
            )
            location_input.click()
            time.sleep(0.5)
            location_input.send_keys(locations[0])
            time.sleep(1)
            try:
                option = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, f"//*[contains(text(),'{locations[0]}')]")
                ))
                option.click()
            except TimeoutException:
                pass
            time.sleep(3)

        perm = check_has_edit_permission(driver)

        if perm == "edit":
            print(f"  ✅ Edit permission confirmed!")
            if is_headless:
                print(f"  💡 Running in the background — no interaction needed.\n")
            else:
                print(f"  💡 Scraping will now run automatically. You can minimize")
                print(f"     this window and keep working on other things.\n")
            scrape_mode = "edit"
        elif perm == "view":
            print(f"  ⚠️  Only 'View Location' found (no edit button)")
            if is_headless:
                print(f"  → Headless mode: auto-selecting read-only mode\n")
                scrape_mode = "readonly"
            else:
                choice = show_no_edit_warning(region)
                if choice == "readonly":
                    scrape_mode = "readonly"
                elif choice == "skip":
                    for s in locations:
                        failed.append({"location": s, "error": "Skipped — no edit permission"})
                    total_time = time.time() - start_time
                    return all_data, failed, total_time, "skip", False
                elif choice == "abort":
                    for s in locations:
                        failed.append({"location": s, "error": "Aborted — no edit permission"})
                    return all_data, failed, 0, "abort", False
        else:
            print(f"  ⚠️  No Edit or View button found on first location")
            if is_headless:
                print(f"  → Headless mode: auto-selecting read-only mode\n")
                scrape_mode = "readonly"
            else:
                choice = show_no_edit_warning(region)
                if choice == "readonly":
                    scrape_mode = "readonly"
                elif choice == "skip":
                    for s in locations:
                        failed.append({"location": s, "error": "Skipped — no edit permission"})
                    total_time = time.time() - start_time
                    return all_data, failed, total_time, "skip", False
                elif choice == "abort":
                    for s in locations:
                        failed.append({"location": s, "error": "Aborted — no edit permission"})
                    return all_data, failed, 0, "abort", False

        permission_checked = True
        driver.get(url)
        time.sleep(5)

    except Exception as e:
        print(f"  ⚠️  Permission check failed: {e}")
        print(f"  → Will detect on first real scrape attempt\n")
        driver.get(url)
        time.sleep(5)

    mode_label = "📝 EDIT" if scrape_mode == "edit" else "👁️ READ-ONLY"
    print(f"  Mode: {mode_label}\n")

    for i, location in enumerate(locations, 1):
        elapsed = time.time() - start_time
        avg = elapsed / max(i - 1, 1)
        remaining = avg * (total - i + 1)
        print(f"  [{i}/{total}] 🏭 {location}  (~{remaining/60:.1f} min left)")

        try:
            if scrape_mode == "readonly":
                row_data = scrape_location_readonly(driver, wait, location, region)
                all_data.append(row_data)
                found = sum(
                    1 for k, v in row_data.items()
                    if k != "Location" and v not in ("", "No Edit Access")
                )
                noedit = sum(1 for v in row_data.values() if v == "No Edit Access")
                print(f"    👁️  Found:{found} NoAccess:{noedit}")
                consecutive_no_edit = 0

            else:
                row_data = scrape_location(driver, wait, location, region)
                all_data.append(row_data)
                a = sum(1 for v in row_data.values() if v == "Active")
                ia = sum(1 for v in row_data.values() if v == "Inactive")
                ua = sum(1 for v in row_data.values() if v == "Unavailable")
                print(f"    ✅ Active:{a} Inactive:{ia} Unavailable:{ua}")
                consecutive_no_edit = 0

        except VPNRequiredError as e:
            print(f"    🔒 {e}")

            if is_headless:
                print(f"    → VPN issue in headless mode. Need visible browser.")
                remaining_locations = locations[i - 1:]
                for s in remaining_locations:
                    failed.append({"location": s, "error": "VPN issue in headless — needs relaunch"})
                total_time = time.time() - start_time
                return all_data, failed, total_time, "needs_relaunch", True

            try:
                import winsound
                winsound.MessageBeep(winsound.MB_ICONEXCLAMATION)
                time.sleep(0.2)
                winsound.MessageBeep(winsound.MB_ICONEXCLAMATION)
            except Exception:
                pass

            vpn_ok = prompt_vpn_connection_browser(driver, url)
            if vpn_ok:
                print(f"  ✅ VPN reconnected! Resuming scrape...")
                print(f"  💡 The browser will continue on its own. Feel free to")
                print(f"     minimize and keep working.\n")

                driver.get(url)
                time.sleep(5)

                if not is_on_actual_site(driver.current_url, region):
                    if is_csa_blocked(driver):
                        vpn_ok2 = prompt_vpn_connection_browser(driver, url)
                        if not vpn_ok2:
                            remaining_locations = locations[i - 1:]
                            for s in remaining_locations:
                                failed.append({"location": s, "error": "VPN disconnected"})
                            break
                    else:
                        if not login_with_browser(driver, region, browser_name):
                            remaining_locations = locations[i - 1:]
                            for s in remaining_locations:
                                failed.append({"location": s, "error": "Re-login failed after VPN reconnect"})
                            break

                try:
                    if scrape_mode == "readonly":
                        row_data = scrape_location_readonly(driver, wait, location, region)
                    else:
                        row_data = scrape_location(driver, wait, location, region)
                    all_data.append(row_data)
                    print(f"    ✅ Retry successful for {location}")
                except Exception as retry_e:
                    print(f"    ❌ Retry failed: {retry_e}")
                    failed.append({"location": location, "error": f"Retry failed: {retry_e}"})

            else:
                print(f"\n  🛑 VPN not reconnected. Saving progress so far...")
                remaining_locations = locations[i - 1:]
                for s in remaining_locations:
                    failed.append({"location": s, "error": "VPN disconnected — user stopped"})
                break

        except NoEditPermissionError as e:
            consecutive_no_edit += 1

            if not permission_checked:
                print(f"    ⚠️  {e}")
                permission_checked = True

                if is_headless:
                    print(f"    → Headless mode: auto-switching to read-only\n")
                    scrape_mode = "readonly"
                    navigate_back_safely(driver, url)
                    try:
                        row_data = scrape_location_readonly(
                            driver, wait, location, region
                        )
                        all_data.append(row_data)
                        found = sum(
                            1 for k, v in row_data.items()
                            if k != "Location" and v not in ("", "No Edit Access")
                        )
                        print(f"    👁️  Read-only: Found {found} fields")
                    except VPNRequiredError as vpn_e:
                        remaining_locations = locations[i - 1:]
                        for s in remaining_locations:
                            failed.append({"location": s, "error": "VPN issue in headless"})
                        return all_data, failed, time.time() - start_time, "needs_relaunch", True
                    except Exception as e2:
                        failed.append({"location": location, "error": str(e2)})
                    continue

                choice = show_no_edit_warning(region)
                if choice == "readonly":
                    scrape_mode = "readonly"
                    print(f"\n  → Switching to read-only mode")
                    print(f"  → Re-scraping {location} in read-only...\n")

                    navigate_back_safely(driver, url)

                    try:
                        row_data = scrape_location_readonly(
                            driver, wait, location, region
                        )
                        all_data.append(row_data)
                        found = sum(
                            1 for k, v in row_data.items()
                            if k != "Location" and v not in ("", "No Edit Access")
                        )
                        print(f"    👁️  Read-only: Found {found} fields")
                    except VPNRequiredError as vpn_e:
                        print(f"    🔒 {vpn_e}")
                        vpn_ok = prompt_vpn_connection_browser(driver, url)
                        if not vpn_ok:
                            remaining_locations = locations[i - 1:]
                            for s in remaining_locations:
                                failed.append({"location": s, "error": "VPN disconnected"})
                            break
                        failed.append({"location": location, "error": str(vpn_e)})
                    except Exception as e2:
                        print(f"    ❌ Read-only also failed: {e2}")
                        failed.append({"location": location, "error": str(e2)})
                    continue

                elif choice == "skip":
                    remaining_locations = locations[i - 1:]
                    for s in remaining_locations:
                        failed.append({
                            "location": s,
                            "error": "Skipped — no edit permission",
                        })
                    break

                elif choice == "abort":
                    remaining_locations = locations[i - 1:]
                    for s in remaining_locations:
                        failed.append({
                            "location": s,
                            "error": "Aborted — no edit permission",
                        })
                    total_time = time.time() - start_time
                    return all_data, failed, total_time, "abort", False

            elif consecutive_no_edit >= 3 and scrape_mode == "edit":
                print(f"    ⚠️  {e}")
                print(f"    ⚠️  3 consecutive 'no edit' failures")
                print(f"    → Auto-switching to read-only mode")
                if not is_headless:
                    print(f"    💡 This will continue automatically. You can minimize")
                    print(f"       this window and keep working on other things.\n")
                scrape_mode = "readonly"

                navigate_back_safely(driver, url)

                try:
                    row_data = scrape_location_readonly(
                        driver, wait, location, region
                    )
                    all_data.append(row_data)
                    print(f"    👁️  Read-only fallback successful")
                except VPNRequiredError as vpn_e:
                    if is_headless:
                        remaining_locations = locations[i - 1:]
                        for s in remaining_locations:
                            failed.append({"location": s, "error": "VPN issue in headless"})
                        return all_data, failed, time.time() - start_time, "needs_relaunch", True
                    print(f"    🔒 {vpn_e}")
                    vpn_ok = prompt_vpn_connection_browser(driver, url)
                    if not vpn_ok:
                        remaining_locations = locations[i - 1:]
                        for s in remaining_locations:
                            failed.append({"location": s, "error": "VPN disconnected"})
                        break
                    failed.append({"location": location, "error": str(vpn_e)})
                except Exception as e2:
                    failed.append({"location": location, "error": str(e2)})

            else:
                print(f"    ⚠️  No edit: {e}")
                failed.append({"location": location, "error": str(e)})
                navigate_back_safely(driver, url)

        except Exception as e:
            print(f"    ❌ Error: {e}")
            failed.append({"location": location, "error": str(e)})
            consecutive_no_edit = 0

            try:
                if is_csa_blocked(driver):
                    if is_headless:
                        remaining_locations = locations[i:]
                        for s in remaining_locations:
                            failed.append({"location": s, "error": "VPN issue in headless"})
                        return all_data, failed, time.time() - start_time, "needs_relaunch", True

                    print(f"    🔒 Looks like VPN disconnected!")

                    try:
                        import winsound
                        winsound.MessageBeep(winsound.MB_ICONEXCLAMATION)
                    except Exception:
                        pass

                    vpn_ok = prompt_vpn_connection_browser(driver, url)
                    if vpn_ok:
                        print(f"  ✅ VPN reconnected! Continuing...")
                        print(f"  💡 Feel free to minimize and keep working.\n")
                        driver.get(url)
                        time.sleep(5)

                        if not is_on_actual_site(driver.current_url, region):
                            if not login_with_browser(driver, region, browser_name):
                                remaining_locations = locations[i:]
                                for s in remaining_locations:
                                    failed.append({"location": s, "error": "Re-login failed"})
                                break
                    else:
                        remaining_locations = locations[i:]
                        for s in remaining_locations:
                            failed.append({"location": s, "error": "VPN disconnected"})
                        break
                else:
                    try:
                        for btn in driver.find_elements(By.TAG_NAME, "button"):
                            if "cancel" in btn.text.strip().lower():
                                btn.click()
                                time.sleep(2)
                                break
                        else:
                            driver.get(url)
                            time.sleep(5)
                    except Exception:
                        driver.get(url)
                        time.sleep(5)
            except Exception:
                driver.get(url)
                time.sleep(5)

        time.sleep(1)

    total_time = time.time() - start_time
    success = total - len(failed)
    print(f"\n  ✅ {region}: {success}/{total}, {total_time/60:.1f} min", end="")
    if scrape_mode == "readonly":
        print(f" (read-only mode)")
    else:
        print()

    return all_data, failed, total_time, scrape_mode, False


# ====================================================================
# MAIN
# ====================================================================
def main():
    na_count = len(REGIONS["NA"]["locations"])
    eu_count = len(REGIONS["EU"]["locations"])

    print("=" * 60)
    print(f"  LargeItem Cluster Transfer Scraper")
    print(f"  NA: {na_count} | EU: {eu_count} | Total: {na_count + eu_count}")
    print(f"  Local: {fmt_local()}")
    print(f"  UTC:   {fmt_utc()}")
    print("=" * 60)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(COOKIES_DIR, exist_ok=True)

    # ══════════════════════════════════════════════════════════════
    # STEP 1: VPN CHECK
    # ══════════════════════════════════════════════════════════════
    vpn_ok = prompt_vpn_before_start()
    if not vpn_ok:
        print(f"\n  ❌ Cannot proceed without network connectivity.")
        print(f"     Connect to Corporate VPN Client and re-run the script.")
        input("\n  Press Enter to exit...")
        return

    # ══════════════════════════════════════════════════════════════
    # STEP 2: BROWSER DETECTION
    # ══════════════════════════════════════════════════════════════
    print(f"\n[SETUP] Detecting browsers...\n")

    installed = get_installed_browsers()
    running = get_running_browsers()

    print(f"  Installed: {', '.join(BROWSER_INFO[b]['name'] for b in installed) if installed else 'None found'}")
    print(f"  Running:   {', '.join(BROWSER_INFO[b]['name'] for b in running) if running else 'None'}")

    if not installed:
        show_no_browser_message()
        input("\n  Press Enter to exit...")
        return

    browser_key, needs_kill = pick_best_browser()

    if browser_key is None:
        print("\n  ❌ No browser available. Exiting.")
        input("\n  Press Enter to exit...")
        return

    browser_name = BROWSER_INFO[browser_key]["name"]

    # ══════════════════════════════════════════════════════════════
    # STEP 3: COOKIE CHECK
    # ══════════════════════════════════════════════════════════════
    na_path = os.path.join(COOKIES_DIR, REGIONS["NA"]["cookies_file"])
    eu_path = os.path.join(COOKIES_DIR, REGIONS["EU"]["cookies_file"])
    na_fresh = os.path.exists(na_path) and (time.time() - os.path.getmtime(na_path) < 43200)
    eu_fresh = os.path.exists(eu_path) and (time.time() - os.path.getmtime(eu_path) < 43200)
    headless = na_fresh and eu_fresh

    if headless:
        print(f"\n  🍪 Fresh cookies found → trying HEADLESS mode 👻")
        print(f"  💡 If cookies are still valid, the browser will run")
        print(f"     invisibly with no interaction needed.")
        print(f"     If cookies expired, it will relaunch visibly for login.\n")
    else:
        print(f"\n  🍪 NA: {'✅ fresh' if na_fresh else '❌ expired/missing'} | EU: {'✅ fresh' if eu_fresh else '❌ expired/missing'}")
        print(f"  🖥️  VISIBLE mode — you'll need to log in once.")
        print(f"     After logging in, the browser works on its own.\n")

    # ══════════════════════════════════════════════════════════════
    # STEP 4: LAUNCH BROWSER
    # ══════════════════════════════════════════════════════════════
    if needs_kill:
        kill_browser(browser_key)
    else:
        kill_stale_driver(browser_key)

    print(f"[SETUP] Launching {browser_name}...")
    driver, actual_key = try_launch_with_fallback(browser_key, headless=headless)

    if not driver:
        print(f"\n  ❌ Could not launch any browser.")
        print(f"  Make sure one of the following is installed:")
        print(f"    • Microsoft Edge   • Google Chrome   • Mozilla Firefox")
        input("\n  Press Enter to exit...")
        return

    actual_name = BROWSER_INFO[actual_key]["name"]
    wait = WebDriverWait(driver, 15)
    is_headless = headless
    print(f"  ✅ {actual_name} launched ({'headless' if is_headless else 'visible'})")

    if is_headless:
        load_cookies(driver, "NA")

    # ══════════════════════════════════════════════════════════════
    # STEP 5: SCRAPE NA
    # ══════════════════════════════════════════════════════════════
    overall_mode = "edit"

    na_data, na_failed, na_time, na_mode, na_needs_relaunch = scrape_region(
        driver, wait, "NA", actual_name, is_headless
    )

    if na_needs_relaunch:
        print(f"\n  🔄 Headless mode can't continue — relaunching visible...")

        for cookie_file in [REGIONS["NA"]["cookies_file"], REGIONS["EU"]["cookies_file"]]:
            cookie_path = os.path.join(COOKIES_DIR, cookie_file)
            if os.path.exists(cookie_path):
                try:
                    os.remove(cookie_path)
                    print(f"  🗑️  Removed stale cookies: {cookie_file}")
                except Exception:
                    pass

        driver, actual_key = relaunch_visible(driver, actual_key)

        if not driver:
            print(f"\n  ❌ Could not relaunch browser.")
            input("\n  Press Enter to exit...")
            return

        actual_name = BROWSER_INFO[actual_key]["name"]
        wait = WebDriverWait(driver, 15)
        is_headless = False

        print(f"\n  🔄 Retrying NA in visible mode...")
        na_data, na_failed, na_time, na_mode, _ = scrape_region(
            driver, wait, "NA", actual_name, is_headless
        )

    # ══════════════════════════════════════════════════════════════
    # STEP 6: SCRAPE EU
    # ══════════════════════════════════════════════════════════════
    if na_mode == "abort":
        print(f"\n  🛑 Aborted by user after NA permission check.")
        print(f"  → Saving whatever was collected...\n")
        eu_data, eu_failed, eu_time, eu_mode = [], [], 0, "skip"
        for s in REGIONS["EU"]["locations"]:
            eu_failed.append({"location": s, "error": "Skipped — user aborted after NA"})

    elif na_mode == "vpn_failed":
        print(f"\n  🔒 NA failed due to VPN. Trying EU anyway...\n")
        eu_data, eu_failed, eu_time, eu_mode, eu_needs_relaunch = scrape_region(
            driver, wait, "EU", actual_name, is_headless
        )

        if eu_needs_relaunch:
            driver, actual_key = relaunch_visible(driver, actual_key)
            if driver:
                actual_name = BROWSER_INFO[actual_key]["name"]
                wait = WebDriverWait(driver, 15)
                is_headless = False
                eu_data, eu_failed, eu_time, eu_mode, _ = scrape_region(
                    driver, wait, "EU", actual_name, is_headless
                )
            else:
                eu_data, eu_failed, eu_time, eu_mode = [], [], 0, "skip"
                for s in REGIONS["EU"]["locations"]:
                    eu_failed.append({"location": s, "error": "Browser relaunch failed"})
    else:
        eu_data, eu_failed, eu_time, eu_mode, eu_needs_relaunch = scrape_region(
            driver, wait, "EU", actual_name, is_headless
        )

        if eu_needs_relaunch:
            driver, actual_key = relaunch_visible(driver, actual_key)
            if driver:
                actual_name = BROWSER_INFO[actual_key]["name"]
                wait = WebDriverWait(driver, 15)
                is_headless = False
                eu_data, eu_failed, eu_time, eu_mode, _ = scrape_region(
                    driver, wait, "EU", actual_name, is_headless
                )
            else:
                eu_data, eu_failed, eu_time, eu_mode = [], [], 0, "skip"
                for s in REGIONS["EU"]["locations"]:
                    eu_failed.append({"location": s, "error": "Browser relaunch failed"})

        if eu_mode == "abort":
            print(f"\n  🛑 Aborted by user after EU permission check.")
            print(f"  → Saving whatever was collected...\n")

    if na_mode == "readonly" or eu_mode == "readonly":
        overall_mode = "readonly"

    # ══════════════════════════════════════════════════════════════
    # STEP 7: CREATE OUTPUT
    # ══════════════════════════════════════════════════════════════
    print(f"\n{'='*60}")
    print(f"  📦 CREATING EXCEL")
    print(f"{'='*60}\n")

    excel_path = os.path.join(OUTPUT_DIR, f"Cluster_Transfer_{TIMESTAMP}.xlsx")
    json_path = os.path.join(OUTPUT_DIR, f"cluster_transfer_{TIMESTAMP}.json")

    create_excel(na_data, eu_data, na_failed, eu_failed, overall_mode, excel_path)
    print(f"  💾 {excel_path}")

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({
            "generated_local": fmt_local(),
            "generated_utc": fmt_utc(),
            "browser": actual_name,
            "na_mode": na_mode,
            "eu_mode": eu_mode,
            "NA": {"data": na_data, "failed": na_failed},
            "EU": {"data": eu_data, "failed": eu_failed},
        }, f, indent=2)

    print(f"  💾 {json_path}")

    # ══════════════════════════════════════════════════════════════
    # STEP 7.5: UPLOAD TO CLOUD_DRIVE
    # ══════════════════════════════════════════════════════════════
    upload_files = [excel_path]
    upload_result = "failed"
    try:
        upload_result = upload_to_cloud_drive(driver, upload_files)
    except Exception as e:
        print(f"  ⚠️  Cloud Drive upload error: {e}")
        print(f"     You can manually upload later.\n")

    # ══════════════════════════════════════════════════════════════
    # STEP 8: COMPLETION
    # ══════════════════════════════════════════════════════════════
    try:
        import winsound
        winsound.MessageBeep(winsound.MB_ICONEXCLAMATION)
        time.sleep(0.3)
        winsound.MessageBeep(winsound.MB_ICONEXCLAMATION)
        time.sleep(0.3)
        winsound.MessageBeep(winsound.MB_OK)
    except Exception:
        pass

    print(f"\n{'='*60}")
    print(f"  🏁 COMPLETE!")
    print(f"{'='*60}")
    print(f"  🌐 Browser: {actual_name}")
    print(f"  🕐 Local: {fmt_local()}")
    print(f"  🕐 UTC:   {fmt_utc()}")
    print(f"  🇺🇸 NA: {na_count - len(na_failed)}/{na_count}", end="")
    if na_mode == "readonly":
        print(f" (read-only)")
    elif na_mode in ("skip", "vpn_failed"):
        print(f" (skipped)")
    else:
        print()
    print(f"  🇪🇺 EU: {eu_count - len(eu_failed)}/{eu_count}", end="")
    if eu_mode == "readonly":
        print(f" (read-only)")
    elif eu_mode in ("skip", "vpn_failed"):
        print(f" (skipped)")
    else:
        print()
    print(f"  ⏱️  {(na_time + eu_time)/60:.1f} minutes")

    if upload_result == "synced":
        print(f"  ☁️  Cloud Drive: Copied to sync folder (auto-uploading)")
    elif upload_result == "api":
        print(f"  ☁️  Cloud Drive: Uploaded via API")
    elif upload_result == "manual":
        print(f"  ☁️  Cloud Drive: Opened for manual drag-and-drop")
    else:
        print(f"  ☁️  Cloud Drive: Upload not completed — do it manually:")
        print(f"     {APS_CONFIG_FOLDER_URL}")

    if overall_mode == "readonly":
        print(f"\n  ℹ️  Some data was collected in read-only mode.")
        print(f"  ℹ️  Cells marked 'No Edit Access' could not be read.")
        print(f"  ℹ️  Request edit permissions for full data next time.")

    if na_mode == "vpn_failed" or eu_mode == "vpn_failed":
        print(f"\n  🔒 Some regions failed due to VPN issues.")
        print(f"     Make sure Corporate VPN Client is connected and try again.")

    if na_failed or eu_failed:
        total_failed = na_failed + eu_failed
        show_count = min(len(total_failed), 10)
        print(f"\n  Failed ({len(total_failed)} total):")
        for f in na_failed[:5]:
            print(f"    ❌ NA - {f['location']}: {f['error']}")
        for f in eu_failed[:5]:
            print(f"    ❌ EU - {f['location']}: {f['error']}")
        if len(total_failed) > show_count:
            print(f"    ... and {len(total_failed) - show_count} more (see Failed Locations tab)")

    print(f"\n  📁 {excel_path}")

    # ── Draft email ──
    any_success = (len(na_data) + len(eu_data)) > 0
    if any_success:
        print(f"\n  📧 Drafting email...")
        email_ok = draft_outlook_email(
            excel_path, na_data, eu_data, na_failed, eu_failed,
            na_mode, eu_mode, overall_mode,
        )
        if not email_ok:
            print(f"  ℹ️  You can still attach the file manually:")
            print(f"      {excel_path}\n")

    try:
        os.startfile(excel_path)
    except Exception:
        pass

    if not is_headless:
        input("\nPress Enter to close browser...")

    try:
        driver.quit()
    except Exception:
        pass


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n\n❌ CRASHED: {e}")
        import traceback
        traceback.print_exc()
    finally:
        input("\nPress Enter to close...")